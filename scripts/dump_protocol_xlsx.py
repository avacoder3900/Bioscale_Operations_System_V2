"""
Dump every sheet of a protocol .xlsx to plain text — cell-by-cell — so Claude
can study the procedures without opening Excel. Throwaway script.
"""
import sys
from pathlib import Path
import openpyxl


def dump_workbook(path: Path) -> str:
    wb = openpyxl.load_workbook(path, data_only=False)
    out = []
    out.append(f"=" * 100)
    out.append(f"FILE: {path.name}")
    out.append(f"=" * 100)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.append("")
        out.append(f"---- SHEET: {sheet_name}  (dim {ws.dimensions}, {ws.max_row} rows x {ws.max_column} cols) ----")
        # Try merged-cell awareness: just list merges, then dump cells
        if ws.merged_cells.ranges:
            out.append(f"Merged ranges: {[str(r) for r in ws.merged_cells.ranges]}")
        for row in ws.iter_rows(values_only=False):
            row_has_content = False
            row_parts = []
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                row_has_content = True
                if isinstance(v, str):
                    v = v.replace("\n", " | ").strip()
                row_parts.append(f"{cell.coordinate}={v}")
            if row_has_content:
                out.append("  " + "    ".join(row_parts))
    return "\n".join(out)


if __name__ == "__main__":
    paths = [Path(p) for p in sys.argv[1:]]
    for p in paths:
        print(dump_workbook(p))
        print()
